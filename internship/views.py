from functools import wraps
from rest_framework import status
from rest_framework.response import Response
from rest_framework.views import APIView
from rest_framework.generics import RetrieveAPIView, ListAPIView, DestroyAPIView
from rest_framework.permissions import IsAuthenticated
from django.http import HttpResponse
from drf_spectacular.utils import extend_schema
import cloudinary.uploader
import logging
import threading
from core.exception_handlers import ErrorEnum, ErrorResponse, response_schemas

from .models import InternshipApplication
from .serializers import InternshipApplicationSerializer
from .signals import internship_application_accepted, internship_application_rejected

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)


class CertificateUploadService:
    """Service class for handling certificate uploads"""
    
    @staticmethod
    def upload_certificate(application_id, certificate_content, certificate_name):
        try:
            upload_result = cloudinary.uploader.upload(
                certificate_content,
                resource_type="raw",
                public_id=certificate_name,
                folder="certificates/",
                overwrite=True
            )
            application = InternshipApplication.objects.get(id=application_id)
            application.certificate = upload_result['secure_url']
            application.save()
            logger.info(f"Certificate uploaded successfully: {upload_result['secure_url']}")
            return True
        except Exception as e:
            logger.error(f"Failed to upload certificate: {e} for application id {application_id}")
            return False


def validate_recipients(func):
    """
    Decorator to validate recipients data in the request
    """
    @wraps(func)
    def wrapper(self, request, *args, **kwargs):
        data = request.data
        recipients = data.get('recipients', [])
        
        if not recipients:
            return ErrorResponse(ErrorEnum.ERR_001, "No recipients provided")
        
        if not isinstance(recipients, list):
            return ErrorResponse(ErrorEnum.ERR_002, "Recipients must be a list")
        
        if not all(isinstance(recipient, dict) and 'name' in recipient and 'email' in recipient 
                  for recipient in recipients):
            return ErrorResponse(ErrorEnum.ERR_003, 
                               "Each recipient must be a dictionary with 'name' and 'email' keys")
        
        if not all(isinstance(recipient['name'], str) and isinstance(recipient['email'], str) 
                  for recipient in recipients):
            return ErrorResponse(ErrorEnum.ERR_004, "Name and email must be strings")
        
        if not all('@' in recipient['email'] for recipient in recipients):
            return ErrorResponse(ErrorEnum.ERR_005, "Email addresses must contain '@'")
            
        return func(self, request, *args, **kwargs)
    return wrapper


class InternshipApplicationBaseView:
    """Base class containing common functionality for internship views"""
    
    serializer_class = InternshipApplicationSerializer
    permission_classes = [IsAuthenticated]  # Default permission
    
    def log_error(self, error, data=None):
        """Helper method for consistent error logging"""
        logger.error(f"Error encountered: {error}")
        if data:
            logger.error(f"Request data: {data}")


class InternshipApplicationView(InternshipApplicationBaseView, APIView):
    """Handle internship application submissions"""

    permission_classes = []  # No auth required for submissions

    @extend_schema(
        tags=['Internship Application'], 
        summary='Submit an internship application',
        description='This endpoint is used to submit internship application'
    )
    def post(self, request, *args, **kwargs):
        # return that registration is closed
        # Registration closed response
        logging.info(f"Registration is closed: {request.data}")
        return Response(
            {"detail": "Registration is closed."}, 
            status=status.HTTP_400_BAD_REQUEST
        )
        serializer = self.serializer_class(data=request.data)
        if not serializer.is_valid():
            self.log_error(serializer.errors, request.data)
            return ErrorResponse(ErrorEnum.ERR_006, serializer.errors)
            
        application = serializer.save()
        certificate = request.FILES.get('certificate')
        
        if certificate:
            thread = threading.Thread(
                target=CertificateUploadService.upload_certificate,
                args=(application.id, certificate.read(), certificate.name)
            )
            thread.start()
            
        return Response(
            {"detail": "Application submitted successfully."}, 
            status=status.HTTP_201_CREATED
        )


class InternshipApplicationListView(InternshipApplicationBaseView, ListAPIView):
    """List all internship applications"""

    permission_classes = []
    
    queryset = InternshipApplication.objects.all()

    @extend_schema(
        tags=['Internship Application'], 
        summary='List all internship applications'
    )
    def get(self, request, *args, **kwargs):
        try:
            queryset = self.get_queryset()
            serializer = self.get_serializer(queryset, many=True)
            
            response_data = {
                "total_applications": queryset.count(),
                "applications": serializer.data
            }
            
            logger.info("All applications retrieved successfully")
            return Response(response_data, status=status.HTTP_200_OK)
            
        except Exception as e:
            self.log_error(e)
            return ErrorResponse(ErrorEnum.ERR_003, str(e))


class InternshipApplicationDetailView(InternshipApplicationBaseView, RetrieveAPIView):
    """Retrieve a single internship application"""
    permission_classes = []
    
    queryset = InternshipApplication.objects.all()

    @extend_schema(
        tags=['Internship Application'], 
        summary='Retrieve an internship application by ID'
    )
    def get(self, request, *args, **kwargs):
        logging.info("Application retrieved successfully")
        return super().get(request, *args, **kwargs)


class InternshipApplicationDeleteView(InternshipApplicationBaseView, DestroyAPIView):
    """Delete an internship application"""
    
    queryset = InternshipApplication.objects.all()

    @extend_schema(
        tags=['Internship Application'], 
        summary='Delete an internship application by ID'
    )
    def delete(self, request, *args, **kwargs):
        try:
            application = self.get_object()
            application.delete()
            return Response(
                {"detail": "Application deleted successfully."}, 
                status=status.HTTP_204_NO_CONTENT
            )
        except Exception as e:
            self.log_error(e)
            return ErrorResponse(ErrorEnum.ERR_009, str(e))


class InternshipApplicationExportView(InternshipApplicationBaseView, APIView):
    """Export applications to Excel format"""

    @extend_schema(
        tags=['Internship Application'],
        summary='Export all internship applications to Excel',
        description='This endpoint exports all internship applications to an Excel file.',
        responses={200: 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'}
    )
    def get(self, request, *args, **kwargs):
        try:
            # Create workbook and worksheet
            wb = Workbook()
            ws = wb.active
            ws.title = "Internship Applications"

            # Define headers and column mapping
            headers = [
                'No', 'Email', 'Full Name', 'Twitter Handle', 'LinkedIn', 'Skill',
                'Experience', 'About Skill', 'Certificate', 'Commitment', 
                'Birthdate', 'Application Date'
            ]
            
            field_mapping = {
                'Email': 'email',
                'Full Name': 'full_name',
                'Twitter Handle': 'twitter_handle',
                'LinkedIn': 'linkedin',
                'Skill': 'skill',
                'Experience': 'experience',
                'About Skill': 'about_skill',
                'Certificate': lambda x: x.certificate.url if x.certificate else '',
                'Commitment': lambda x: 'Yes' if x.commitment else 'No',
                'Birthdate': 'birthdate',
                'Application Date': lambda x: str(x.date_created)
            }

            # Write headers
            for col_num, header in enumerate(headers, 1):
                ws[f'{get_column_letter(col_num)}1'] = header

            # Write data
            applications = InternshipApplication.objects.all()
            for row_num, application in enumerate(applications, 2):
                ws[f'A{row_num}'] = row_num - 1  # Numbering
                
                for col_num, header in enumerate(headers[1:], 2):  # Skip 'No' column
                    field = field_mapping[header]
                    value = field(application) if callable(field) else getattr(application, field)
                    ws[f'{get_column_letter(col_num)}{row_num}'] = value

            # Prepare response
            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = 'attachment; filename=internship_applications.xlsx'
            wb.save(response)

            return response
            
        except Exception as e:
            self.log_error(e)
            return ErrorResponse(ErrorEnum.ERR_010, str(e))


class BaseInternshipEmailView(InternshipApplicationBaseView, APIView):
    """
    Base view for sending internship emails
    """
    permission_classes = []
    signal = None
    email_subject = ''
    email_template = ''
    email_attachments = []
    
    @extend_schema(
        tags=['Internship Application'],
        responses={200: 'Email sent successfully'}
    )
    @validate_recipients
    def post(self, request, *args, **kwargs):
        try:
            recipients = request.data.get('recipients', [])
            
            for recipient in recipients:
                self.signal.send(
                    sender=self.__class__,
                    name=recipient.get('name'),
                    email=recipient.get('email'),
                    subject=self.email_subject,
                    template=self.email_template,
                    attachments=self.email_attachments
                )
                
            return Response(
                {'detail': 'Email sent successfully'}, 
                status=status.HTTP_200_OK
            )
        except Exception as e:
            self.log_error(e)
            return ErrorResponse(ErrorEnum.ERR_011, str(e))


class InternshipApplicationSendMailView(BaseInternshipEmailView):
    """
    View for sending acceptance emails to intern applicants
    """
    signal = internship_application_accepted
    email_subject = 'Welcome to TCU Cohort 3.0 – Internship Onboarding Information'
    email_template = 'email/acceptance.html'
    email_attachments = ['files/ProjectTimelines.pdf', 'files/TermsAndConditions.pdf']
    
    @extend_schema(
        summary='Send email to accepted Intern applications',
        description='This endpoint sends an acceptance email to selected Intern applications.',
    )
    def post(self, request, *args, **kwargs):
        return super().post(request, *args, **kwargs)


class InternshipApplicationRejectionMailView(BaseInternshipEmailView):
    """
    View for sending rejection emails to intern applicants
    """
    signal = internship_application_rejected
    email_subject = 'TCU Cohort 3.0 – Application Update'
    email_template = 'email/rejection.html'
    email_attachments = []
    
    @extend_schema(
        summary='Send rejection email to Intern applications',
        description='This endpoint sends a rejection email to Intern applications.',
    )
    def post(self, request, *args, **kwargs):
        return super().post(request, *args, **kwargs)